import openpyxl
from commom import setting
from commom.mylog import Log
from selenium.webdriver.common.by import By
class ReadExcel():
    def read_data(self,sheet,file=setting.testdatadir):
        global worksheet
        try:
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook[sheet]
        except Exception as e:
            Log.error("出现以下异常{}".format(e))
        rows=worksheet.max_row
        columns=worksheet.max_column
        title=[]
        data=[]
        for i in range(1,columns+1):
            title.append(worksheet.cell(1,i).value)
        for j in range(2,rows+1):
            data_real={title[0]:worksheet.cell(j,1).value,title[1]:worksheet.cell(j,2).value,title[2]:worksheet.cell(j,3).value,title[3]:worksheet.cell(j,4).value,title[4]:worksheet.cell(j,5).value,title[5]:worksheet.cell(j,6).value,title[6]:worksheet.cell(j,7).value}
            data.append(data_real)
        return data
if __name__ == '__main__':
    a=ReadExcel().read_data("字段管理")
    b=a[1]["param"]
    c=eval(b)
    print(c,type(c))
    print(len(c))
    for i in range(len(c)):
        print(c[i])
    #print(c,type(c))
    #c=a[4]["action"]

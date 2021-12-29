import openpyxl
from common import setting
from common.mylog import Log
from selenium.webdriver.common.by import By
class ReadExcel():
    def read_data(self,sheet,file=setting.testdatadir):
        global worksheet
        try:
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook[sheet]
        except Exception as e:
            Log.error("出现以下异常{}".format(e))
        rows=worksheet.max_row
        columns=worksheet.max_column
        title=[]
        data=[]
        for i in range(1,columns+1):
            title.append(worksheet.cell(1,i).value)
        for j in range(2,rows+1):
            data_real={title[0]:worksheet.cell(j,1).value,title[1]:worksheet.cell(j,2).value,title[2]:worksheet.cell(j,3).value,title[3]:worksheet.cell(j,4).value,title[4]:worksheet.cell(j,5).value,title[5]:worksheet.cell(j,6).value,title[6]:worksheet.cell(j,7).value}
            data.append(data_real)
        return data
if __name__ == '__main__':
    a=ReadExcel().read_data("批次管理")
    #print(c,type(c))
    #c=a[4]["action"]


"""
使用xlrd
"""
"""import xlrd
from common import setting
from common.mylog import Log
from selenium.webdriver.common.by import By
class ReadExcel():
    def read_data(self,sheet,file=setting.testdatadir):
        global worksheet
        try:
            workbook = xlrd.open_workbook(file)
            worksheet = workbook.sheet_by_name(sheet)
        except Exception as e:
            Log.error("出现以下异常{}".format(e))
        rows=worksheet.nrows
        columns=worksheet.ncols
        title = worksheet.row_values(0)
        print(rows,columns,title)
        data=[]
        for i in range(1,rows):
            everydata=dict(zip(title,worksheet.row_values(i)))
            data.append(everydata)
        print(data)
        return data
if __name__ == '__main__':
    a=ReadExcel().read_data("批次管理")"""
